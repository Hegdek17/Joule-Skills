#!/usr/bin/env python3
"""
generate_excel.py — Generates a formatted Excel timesheet report from presales-log.json

Usage:
  python3 generate_excel.py \
    --log presales-log.json \
    --start 2026-07-21 \
    --end 2026-07-27 \
    --output presales-report-2026-W30.xlsx
"""

import argparse
import json
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="Generate weekly timesheet Excel report")
    parser.add_argument("--log", required=True, help="Path to presales-log.json")
    parser.add_argument("--start", required=True, help="Start date YYYY-MM-DD (inclusive)")
    parser.add_argument("--end", required=True, help="End date YYYY-MM-DD (inclusive)")
    parser.add_argument("--output", required=True, help="Output .xlsx file path")
    args = parser.parse_args()

    # Load log
    try:
        with open(args.log, "r") as f:
            entries = json.load(f)
    except FileNotFoundError:
        print(f"ERROR: Log file not found: {args.log}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"ERROR: Invalid JSON in log file: {e}")
        sys.exit(1)

    try:
        start_date = datetime.strptime(args.start, "%Y-%m-%d").date()
        end_date = datetime.strptime(args.end, "%Y-%m-%d").date()
    except ValueError as e:
        print(f"ERROR: Invalid date format: {e}")
        sys.exit(1)

    # Filter entries for the week
    week_entries = []
    for entry in entries:
        try:
            entry_date = datetime.strptime(entry["date"], "%Y-%m-%d").date()
            if start_date <= entry_date <= end_date:
                week_entries.append(entry)
        except (KeyError, ValueError):
            continue

    # Sort by date, then customer name
    week_entries.sort(key=lambda x: (x.get("date", ""), x.get("customer_name", "").lower()))

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    week_label = f"{start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Pre-Sales Activity Timesheet  |  {week_label}"
    title_cell.font = Font(bold=True, size=13, color="1F4E79")
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 28

    ws.append([])  # blank spacer row

    # Header row
    headers = ["Date", "Day", "Customer Name", "Hours", "Type", "Opportunity No.", "Notes"]
    ws.append(headers)
    header_row_num = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_num, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[header_row_num].height = 20

    # Data rows
    total_hours = 0.0
    cf_hours = 0.0
    ncf_hours = 0.0

    for i, entry in enumerate(week_entries):
        try:
            entry_date = datetime.strptime(entry["date"], "%Y-%m-%d").date()
        except (KeyError, ValueError):
            continue

        day_name = entry_date.strftime("%A")
        hours = float(entry.get("hours", 0))
        activity = entry.get("activity_type", "")

        row_data = [
            entry_date.strftime("%d %b %Y"),
            day_name,
            entry.get("customer_name", ""),
            hours,
            activity,
            entry.get("opportunity_number", ""),
            entry.get("notes", ""),
        ]
        ws.append(row_data)
        data_row_num = ws.max_row
        row_fill = alt_fill if i % 2 == 1 else None

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row_num, column=col)
            cell.border = thin_border
            cell.alignment = left_align if col in (3, 7) else center_align
            if row_fill:
                cell.fill = row_fill

        total_hours += hours
        if activity == "CF":
            cf_hours += hours
        elif activity == "NCF":
            ncf_hours += hours

    # Blank spacer before totals
    ws.append([])

    # Totals row
    ws.append([
        "",
        "",
        "TOTAL",
        total_hours,
        f"CF: {cf_hours:.1f}  |  NCF: {ncf_hours:.1f}",
        "",
        f"{len(week_entries)} entries"
    ])
    total_row_num = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_num, column=col)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[total_row_num].height = 20

    # Set column widths
    col_widths = [14, 12, 32, 8, 12, 18, 42]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header row
    ws.freeze_panes = ws.cell(row=header_row_num + 1, column=1)

    # Save
    try:
        wb.save(args.output)
    except Exception as e:
        print(f"ERROR: Could not save file: {e}")
        sys.exit(1)

    print(f"OK: Report saved to {args.output}")
    print(f"    Entries: {len(week_entries)} | Total: {total_hours:.1f} hrs | CF: {cf_hours:.1f} | NCF: {ncf_hours:.1f}")


if __name__ == "__main__":
    main()
